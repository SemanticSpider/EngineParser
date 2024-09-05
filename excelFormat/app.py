from kivymd.app import MDApp
from kivy.lang import Builder
from kivy.core.window import Window
from kivy.animation import Animation
from kivymd.uix.screen import MDScreen
from kivymd.uix.card import MDCard
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.floatlayout import MDFloatLayout
from kivymd.uix.dialog import (MDDialog, MDDialogHeadlineText, MDDialogContentContainer, MDDialogButtonContainer)
from kivymd.uix.progressindicator import MDLinearProgressIndicator
from kivy.properties import StringProperty, NumericProperty, ObjectProperty
from kivy.uix.button import Button
from kivymd.uix.button import MDButton
from plyer import filechooser
import os
import pandas as pd
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
import re
import json
import sys
import requests
from time import sleep
import random
from random import uniform
from bs4 import BeautifulSoup
import threading
import asyncio


class LoadCard(MDCard):

    loadFile = StringProperty()
    label = StringProperty()
    cardColor = ObjectProperty()

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.theme_cls.theme_style_switch_animation = True
        self.theme_cls.theme_style_switch_animation_duration = 0.5
        Window.bind(on_drop_file=self.selectFile)
        
    # Метод, вызываемый при нажатии на кнопку загрузки файла
    def changeFile(self):        
        # Выбор фалйа в filechooser
        choose = filechooser.open_file(on_selection=self.selectFile)

    # Обработчик события выбора файла в filechooser, записывает путь к выбранному файлу в поле loadFile
    def selectFile(self, *args):
        # Получаем путь к файлу
        if(type(args[0]) == list):
            # Из filechooser
            self.loadFile = args[0][0] if len(args[0]) != 0 else self.loadFile
        else:
            # Из Drag-n-Drop
            self.loadFile = args[1].decode(encoding="utf-8")
        
        # Осуществляем проверку Excel расширения
        if self.loadFile and self.loadFile.endswith((".xls", ".xlt", ".xlsx", ".xlsm", ".xltx", ".xltm")):
            self.cardColor = [0.0, 0.5019607843137255, 0.0, 0.3]
            self.label = "Файл {} загружен!".format(self.loadFile.split('\\').pop())
        elif self.loadFile == "":
            pass        
        else: 
            self.cardColor = [1.0, 0.0, 0.0, 0.5]
            self.label = "Принимаются только файлы excel!"
            self.loadFile = ""

    # Ограничиваем дефолтное поведение карточки, чтобы убрать автоматическое изменение цвета
    def set_properties_widget(self):
        super().set_properties_widget()
        self.md_bg_color = self.cardColor
        return True
    
class LoadDirectory(MDBoxLayout):

    savePath = StringProperty()

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        with open(os.path.dirname(os.path.realpath(__file__)) + "\\loadDirectory.txt", "a+") as loadDir:
            loadDir.seek(0, 0)
            self.savePath = loadDir.read()
    
    def chooseDirectory(self):

        def selectDirectory(selectedDirectory):
            self.savePath = selectedDirectory = selectedDirectory[0] if(len(selectedDirectory)) != 0 else self.savePath
            with open(os.path.dirname(os.path.realpath(__file__)) + "\\loadDirectory.txt", "w") as loadDir:
                loadDir.write(self.savePath)

        directory = filechooser.choose_dir(on_selection=selectDirectory)

class LaunchButton(MDButton):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class LoaderDialog(MDDialog):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)


# Класс окна
class Root(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class EngineParser(MDApp):

    loadProgress = NumericProperty(0)

    # Дефолтные настройки приложения
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        Window.maximize()
 
    def build(self):
        self.loader = MDDialog(
            MDDialogHeadlineText(
                text="Сбор информации может занять время..."
            ),
            MDDialogContentContainer(
                MDLinearProgressIndicator(
                    size_hint_x=.5,
                    pos_hint= {'center_x': .5, 'center_y': .5},
                    value=0,
                    type="indeterminate",
                    running_indeterminate_duration=5,
                    theme_bg_color="Custom",
                    track_color=[0.4980, 0.4980, 0.4980, 0.4],
                ),
            )
        )
        self.theme_cls.primary_palette = "Green"
        self.screen = Root()

        loadCard = LoadCard(loadFile = "", label="Загрузите файл", cardColor=self.theme_cls.backgroundColor)        
        loadDirectory = LoadDirectory()
        launchButton = LaunchButton()

        self.screen.add_widget(loadCard)
        self.screen.add_widget(loadDirectory)
        self.screen.add_widget(launchButton)
        # print([widget for widget in self.screen.children if widget.name == "LoadCard"][0].buttonText)
        return self.screen
    
    def launchParse(self, **kwargs):

        def parse(args):
            self = args["self"]
            goodsExcelInfo = self.excelFormat(args["loadFile"])
            benefitProducts = asyncio.run(self.searchBenefit(goodsExcelInfo, args["dialog"]))
            self.createExcel(masToPush=benefitProducts, pathToSave=args["savePath"], fileName=args["loadFile"].split('\\').pop())
            args["dialog"].dismiss()

        properties = {widget.name: widget for widget in self.screen.children if widget.name in ["LoadCard", "LoadDirectory"]}
        loadFile, savePath = [properties["LoadCard"].loadFile, properties["LoadDirectory"].savePath]

        if(loadFile):
            

            # loader = LoaderDialog()
            self.loader.open()
            self.loader.auto_dismiss = False
            thread = threading.Thread(target=parse, args=({"self": self, "loadFile": loadFile, "savePath": savePath, "dialog":  self.loader},), daemon=True)
            thread.start()
    
    def excelFormat(self, excelFile):
        # Читаем с openpyxl
        wb = load_workbook(excelFile, data_only=True)
        sheet = wb.active

        df = pd.DataFrame(sheet.values)

        artikulMas, name, col, price, cost = [pd.Series()]*5

        dataDeletePattern = re.compile(r"(?:(?<=[ ])\d{2}[\.-]\d{2}[\.-]\d{4}\b)|(?:(?<=[ ])\d{4}[\.-]\d{2}[\.-]\d{2}[ :-]?(?:\d{2}[ :]\d{2}[ :]\d{2})?\b)|(?:(?<=[ ])\b(?:\d{2}[\.-]){2}[А-ЯA-Z]+\b)")
        # artikulSearch = re.compile(r"(?:\b(?:\d+[ ])+\d+\b)|(?:\b[A-Z]+[\. -]?(?:\d+[\.-]?)+\b)|(?:\b(?<![\.,])(?<!ГОСТ )(?:[A-ZА-Я\d]+[\.-])+[A-ZА-Я\d]+\b)|(?:\b\d+[A-ZА-Я]+\d+\b)|(?:\b(?<!ГОСТ )\d{5,}[A-ZА-Я]+\b)|(?:\b(?<!ГОСТ )\d{7,}\b)|(?:\b(?:[A-ZА-Я]\([A-ZА-Я]\)-\d+)\b)")
        artikulSearch = re.compile(r"(?:\b(?<![\.,])(?<!ГОСТ )(?:[A-ZА-Я\d]+[\.-])+[A-ZА-Я\d]+\b)|(?:\b(?:\d+[ ])+\d+\b)|(?:\b[A-Z]+[\. -]?(?:\d+[\.-]?)+\b)|(?:\b\d+[A-ZА-Я]+\d+\b)|(?:\b(?<!ГОСТ )\d{5,}[A-ZА-Я]+\b)|(?:\b(?<!ГОСТ )\d{7,}\b)|(?:\b(?:[A-ZА-Я]\([A-ZА-Я]\)-\d+)\b)")
        

        while(df.iloc[0].isna().sum() == df.iloc[0].size):
            df = df.drop(df.index[0])
        for index, column in df.items():
            if(column.isna().sum() == column.size): 
                df = df.drop(index, axis=1)
                continue

            if(column.astype(str).str.contains(dataDeletePattern).any()): 
                df[index], column = [column.astype(str).str.replace(dataDeletePattern, "", regex=True)]*2

            if(column.astype(str).str.contains(r"(?:[Аа]ртикул)|(?:[Кк]аталожный номер)").any()): artikulMas = column.dropna()

            # if(column.astype(str).str.contains(r"(?:[Аа]ртикул)|(?:[Кк]аталожный номер)").any()): artikulMas = artikulMas.combine(column, lambda s1, s2: s1 if s2 == None else s2).dropna()
            # elif(column.astype(str).str.contains(artikulSearch).any()):
            #     artikulMas = column.dropna() if artikulMas.empty else artikulMas.combine(column, lambda s1, s2: s2 if s1 == None else s1).dropna()

            # if(column.astype(str).str.contains(artikulSearch).any()): artikulMas = artikulMas.combine_first(column.dropna())
            # if(column.astype(str).str.contains(r"(?<!\d{2}.)\d{2,}[\.-]\d{3,}(([\.-]\d+)+)?").any()): artikulMas.append(column.dropna())
            if(column.astype(str).str.contains(r"(?:[Нн]аименование)|(?:ТМЦ)").any()): name = column.dropna()
            elif(column.astype(str).str.contains(r"[Кк]ол-во").any()): col = column.dropna()
            elif(column.astype(str).str.contains(r"[Цц]ена").any()): price = column.dropna()
            elif(column.astype(str).str.contains(r"(?:[Сс]тоимость)|([Сс]умма)").any()): cost = column.dropna()
            
        resultObject = {
            "goods": [],
            "errors": []
        }  

        for index, goodsName in name.items():
            if index in artikulMas.index:
                print(artikulMas.loc[index], " => ", re.search(artikulSearch, str(artikulMas.loc[index])))
                search = re.split(r"/", str(artikulMas.loc[index])) if re.search(artikulSearch, str(artikulMas.loc[index])) else []
            else:
                search = re.findall(artikulSearch, str(goodsName))
                
            if len(search) != 0:
                mainArtikul = search[0]
                if len(search) > 1:
                    for artikul in search:
                        if len(re.sub(r"[ \.-]", "", artikul)) > len(re.sub(r"[ \.-]", "", mainArtikul)): mainArtikul = artikul
                resultObject["goods"].append({
                    "mainArtikul": mainArtikul,
                    "artikuls": search,
                    "name": name.loc[index] if not name.empty and index in name.index else None,
                    "amount": col.loc[index] if not col.empty and index in name.index else None,
                    "price": price.loc[index] if not price.empty and index in name.index else None,
                    "cost": cost.loc[index] if not cost.empty and index in name.index else None,
                })
            else: resultObject["errors"].append(goodsName)
        return resultObject
    
    # Создание итоготов таблицы
    def createExcel(self, masToPush=[], pathToSave="", fileName=""):
        try:
            # Создание рабочей книги и выбор листа
            wb = op.Workbook()
            sheet = wb.active

            # Столбцы заголовков
            title = [
                "Название",
                "Название из экселя",
                "Артикул",
                "Артикул из экселя",
                "Цена",
                "Количество",
                "Количество из экселя",
                "Источник", 
                "Бренд"
            ]

            # Параметры рамки
            bd = Side(style="thin", color="000000")

            # Итератор, создающий сттроку заголовков
            row, column = 1, 1
            for key in title:
                letter = sheet.cell(row=row, column=column).column_letter
                sheet.cell(row=row, column=column).value = key
                sheet.cell(row=row, column=column).font = Font(size=12, bold=True)
                sheet.cell(row=row, column=column).border = Border(left=bd, right=bd, bottom=bd, top=bd)
                sheet.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                sheet.column_dimensions[letter].width = 30
                column += 1

            # Заполнение таблицы данными
            row, column = 2, 1
            while len(masToPush) != 0:
                current = masToPush.pop(0)
                if current["Количество"] == "Товар не найден" or int(current["Количество"]) < int(current["Количество из экселя"]): 
                    fill = PatternFill("solid", fgColor="FF3333")
                else: 
                    fill = PatternFill("none")
                for key in title:
                    letter = sheet.cell(row=row, column=column).column_letter
                    sheet.cell(row=row, column=column).fill = fill
                    sheet.cell(row=row, column=column).border = Border(left=bd, right=bd, bottom=bd, top=bd)
                    
                    # if len(str(current[key])) > widthObject[letter]:
                    #     widthObject[letter] = len(str(current[key])) + 10
                    #     sheet.column_dimensions[letter].width = widthObject[letter]
                    sheet.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    sheet.cell(row=row, column=column).value = current[key]
                    column += 1
                row += 1
                column = 1
            # Сохранение таблицы
            wb.save(pathToSave + "\\" + fileName.replace(".xlsx", "_результат.xlsx"))
            return True
        except PermissionError:
            print("Перед записью закройте эксель")
            return False
    
    async def get_product_info(self, article, productName):

        url = f"https://www.autoopt.ru/search/index?search={article}"
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
        }

        req = requests.get(url, headers=headers)

        src = req.text

        result_mas = []

        soup = BeautifulSoup(src, "lxml")

        all_products = soup.find_all(
            "div", class_="n-catalog-item relative grid-item n-catalog-item__product"
        )

        if len(all_products) == 0:
            analogLink = soup.find(
                "table", class_="table light-head search-table search-analogs"
            )
            if analogLink:
                analogLink = analogLink.find(
                    "a", class_="js-anchor-tab"
                )["href"]
                url =  f"https://www.autoopt.ru{analogLink}"
                req = requests.get(url, headers=headers)
                src = req.text
                soup = BeautifulSoup(src, "lxml")
                all_products = soup.find_all(
                    "div", class_="n-catalog-item relative grid-item n-catalog-item__product"
                )
        
        if len(all_products) != 0:
            mainProduct = None
            for product in all_products:
                if product.find("a", class_="n-catalog-item__name-link").text.strip() == productName : mainProduct = product
            if mainProduct is None: mainProduct = all_products[0]

            # получение кода товара
            product_code = mainProduct.find(
                "div", class_="n-catalog-item__photo-code"
            ).find(
                "span", class_="string bold n-catalog-item__click-copy n-catalog-item__ellipsis"
            ).text

            # получение названия бренда товара
            product_brand = mainProduct.find(
                "div", class_="n-catalog-item__brand d-none d-md-table-cell"
            ).find(
                class_="n-catalog-item__ellipsis"
            ).text.strip()

            # получение артикула товара
            product_article = mainProduct.find(
                "div", class_="n-catalog-item__article"
            ).find(
                "span", class_="n-catalog-item__ellipsis"
            ).text

            # получение названия товара
            product_name = mainProduct.find(
                "div", class_="string"
            ).find(
                "a", class_="n-catalog-item__name-link"
            ).text.strip()

            supplier_price =  mainProduct.find("supplier-price")
            if supplier_price: supplier_price = json.loads(supplier_price.get(":offer"))

            # получение цен товара с разными днями доставки
            if mainProduct.find("offers") == None:
                pass
            else:
                product_prices = mainProduct.find(
                    "offers")            
                product_price_obj = [{"Цена": mainProduct["price"], "Количество": mainProduct["quantity"], "Доставка": mainProduct["deliveryDate"]} for mainProduct in json.loads(product_prices[":offers"])]
            
            # получение цен товара
            if mainProduct.find("div", class_="n-catalog-item__price-box col-12 col-md pr-0 mb-2") == None:
                pass
            else:
                product_price_obj = {}

                product_prices_div = mainProduct.find(
                    "div", class_="n-catalog-item__price-box col-12 col-md pr-0 mb-2"
                )

                if supplier_price: product_price_obj["Опт 3"] = supplier_price["price"]
                else:
                    product_prices = product_prices_div.find(
                        "ul"
                    ).find_all(
                        "li"
                    )
                    for price in product_prices:
                        type_price = price.find(
                            class_="fake mr-2 link-color price5").text.strip()
                        cur_price = price.find(class_="gray").text
                        product_price_obj[type_price] = re.search(r"\d+(?:[.]\d+)?", cur_price).group(0)

            # получение количества товаров
            count = mainProduct.find(
                "span", class_="fake grass bold mr-0"
            )

            if count != None:
                product_count = mainProduct.find(
                    "div", class_="n-catalog-item__count-box"
                ).find(
                    "span", class_="fake grass bold mr-0"
                ).text.strip()[:-1]
                product_count = int(product_count)
            else:
                count = mainProduct.find(
                    "span", class_="fake link-gray"
                )
                if count != None:
                    product_count = count.text
                else:
                    offers = mainProduct.find(
                        "offers"
                    )
                    if offers != None:
                        product_count = offers.get(":offers")
                        product_count = json.loads(product_count)[0]["quantity"]
                    elif mainProduct.find("supplier-price") != None:
                        # print("SUPER PRODUCT COUNT", product_count.text)
                        product_count = supplier_price["quantity"]
                        


            product_info = {
                "Код товара": product_code,
                "Название": product_name,
                "Название": product_name,
                "Бренд товара": product_brand,
                "Артикул товара": product_article,
                "Цены товара": product_price_obj,
                "Количество на складе": product_count,
            }

            result_mas.append(product_info)
            

            first_product_url = mainProduct.find(
                "div", class_="string"
            ).find(
                "a", class_="n-catalog-item__name-link"
            )

            first_product_url = first_product_url.get("href")
            # Если есть ссылка на детальную страницу
            if first_product_url:  

                detail_url = f"https://www.autoopt.ru{first_product_url}"

                req = requests.get(detail_url, headers=headers)

                src = req.text

                detail_page = BeautifulSoup(src, "lxml")
                
                if  detail_page.find("div", class_="table-responsive-md analogs-container"):
                    analog_mas = detail_page.find(
                        "div", class_="table-responsive-md analogs-container"
                    ).find_all(
                        "div", class_="n-catalog-item n-catalog-item__product-item relative grid-item"
                    )

                    for analog in analog_mas:
                        # получение кода товара
                        analog_code = analog.find(
                            "div", class_="n-catalog-item__photo-code"
                        ).find(
                            "span", class_="string bold n-catalog-item__click-copy"
                        ).text

                        # получение названия бренда товара
                        analog_brand = analog.find(
                            "div", class_="n-catalog-item__brand d-md-table-cell n-catalog-clear"
                        ).find(
                            "div", class_="d-md-none__search"
                        )

                        analog_brand = analog_brand.find("a").text.strip() if analog_brand.find("a") else analog_brand.text.strip()

                        # получение артикула товара
                        analog_article_container = analog.find(
                            "div", class_="n-catalog-item__article"
                        )
                        analog_article = analog_article_container.find(
                            "span", class_="string bold nowrap n-catalog-item__click-copy"
                        ).text + ", " + analog_article_container.find(
                            "span", class_="string nowrap n-catalog-item__click-copy n-catalog-item__articles"
                        ).text

                        # получение названия товара
                        analog_name = analog.find(
                            "div", class_="n-catalog-item__name"
                        ).find(
                            "a", class_="n-catalog-item__name-link actions name-popover"
                        ).text.strip()


                        # получение цен товара с разными днями доставки
                        if analog.find("offers") == None:
                            pass
                        else:
                            analog_prices = analog.find(
                                "offers")            
                            analog_price_obj = [{"Цена": price["price"], "Количество": price["quantity"], "Доставка": price["deliveryDate"]} for price in json.loads(analog_prices[":offers"])]
                        
                        # получение цен товара
                        if analog.find("div", class_="n-catalog-item__price-box col-12 col-md pr-0 mb-2") == None:
                            pass
                        else:
                            analog_price_obj = {}
                            analog_prices = analog.find(
                                "div", class_="n-catalog-item__price-box col-12 col-md pr-0 mb-2"
                            ).find(
                                "ul"
                            ).find_all(
                                "li"
                            )



                            for price in analog_prices:
                                type_price = price.find(
                                    class_="fake mr-2 link-color price3").text.strip()
                                cur_price = price.find(class_="gray").text
                                analog_price_obj[type_price] = re.search(r"\d+(?:[.]\d+)?", cur_price).group(0)

                        # получение количества товаров
                        count = analog.find(
                            "span", class_="fake grass bold mr-0"
                        )

                        if count != None:
                            analog_count = analog.find(
                                "div", class_="n-catalog-item__count-box"
                            ).find(
                                "span", class_="fake grass bold mr-0"
                            ).text.strip()[:-1]
                            analog_count = int(analog_count)
                        else:
                            count = analog.find(
                                "span", class_="fake link-gray"
                            )
                            if count != None:
                                analog_count = count.text
                            else:
                                analog_count = analog.find(
                                    "offers"
                                ).get(":offers")
                                analog_count = json.loads(analog_count)[0]["quantity"]


                        analog_info = {
                            "Код товара": analog_code,
                            "Название": analog_name,
                            "Бренд товара": analog_brand,
                            "Артикул товара": analog_article,
                            "Цены товара": analog_price_obj,
                            "Количество на складе": analog_count,
                        }

                        result_mas.append(analog_info)
        await asyncio.sleep(0)
        return result_mas

    async def get_info_autopiter(self, article):
            
        # Список прокси
        proxies = [
            "http://JI2BQ8T6G7:dqSy5wPIKv@103.82.103.8:40253",
            "http://qrX2br:0SrrNc@88.218.75.218:9367",
            "http://qrX2br:0SrrNc@88.218.72.250:9833"
        ]

        # Заголовки для запроса
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            # "Set-Cookie": "guestId=XOuTXkQ01GTNSUl3c_p3f; Max-Age=2592000;"
        }

        # Функция для получения случайного прокси
        def get_proxy(proxies):
            proxy = random.choice(proxies)
            return {"http": proxy, "https": proxy}

        def get_href_product(article):

            url = f"https://autopiter.ru/goods/{article}"
            print(article)
        
            proxy = get_proxy(proxies)
            # print(proxy)

            req = requests.get(url, headers=headers)

            src = req.text
            
            soup = BeautifulSoup(src, "lxml")

            product = soup.find("div", "IndividualTableRow__row___111l8")

            if product:                    
                product_href = product.find(
                    "div", class_="IndividualTableRow__numberColumn___36MQf"
                ).find("a", class_="IndividualTableRow__numberLink___1-eq1 common__link___1TmCU").get("href")
            
            else: return None
            return product_href
        
        def get_product_info_autopiter(href):

            url_product = f"https://autopiter.ru{href}"

            req_product = requests.get(url_product, headers=headers)

            src_product = req_product.text

            soup_product = BeautifulSoup(src_product, "lxml")
            
            table_info = soup_product.find(
                "table", class_="NonRetailAppraiseTable__table___7gnpi"
            )


            # Если таблица с информацией не найдена
            if table_info == None:
                return {}

            products_information = table_info.find("tbody").find_all(
                "tr", class_="NonRetailAppraiseTR__tr___3rq83"
            )

            full_product_info = []

            for product_info in products_information:
                
                # Получение региона продукта
                product_region = product_info.find(
                    "div", class_="NonRetailAppraiseTR__regionInner___3FAjj"
                ).text

                # Получение бренда продукта
                product_brand = product_info.find(
                    "span", class_="NonRetailAppraiseTR__brandLink___v2r3Y ModalButton__root___pxLdy"
                ).find(
                    "span", class_="ModalButton__button___3fjTP new-common__unstyledButton___zigzA"
                ).text

                # Получение артикула продукта
                product_article = product_info.find_all("td")[2].text

                # Получение имени продукта
                product_name = product_info.find(
                    "td", class_="NonRetailAppraiseTR__nameCell___2k4_I"
                ).find("div").text

                # Получение количетсва продукта
                product_count = product_info.find(
                    "td", class_="NonRetailAppraiseTR__quantityCell___nzLgM"
                ).find(
                    "div", class_="NonRetailAppraiseTR__quantity___18gVh"
                ).text

                # Получение информации через сколько придет продукт
                product_time = product_info.find(
                    "td", class_="NonRetailAppraiseTR__daysCell___YR7jI"
                ).find("span", "NonRetailAppraiseTR__days___UfW_V").text

                # Получение стоимости продукта
                product_price = product_info.find(
                    "td", "NonRetailAppraiseTR__priceCell___2hvw7"
                ).find("div", "NonRetailAppraiseTR__priceWrapper___2sj3p").text


                full_product_info.append(
                    {
                        "Регион поставщика": product_region,
                        "Производитель": product_brand,
                        "Номер": product_article,
                        "Наименование": product_name,
                        "Наличие, шт": product_count,
                        "Срок (дн)": product_time,
                        "Цена": product_price,
                    }
                )

            return full_product_info
        
        href = get_href_product(article)    
        await asyncio.sleep(0)

        if href:
            product_info = get_product_info_autopiter(href)
            if product_info: return product_info

        return []
    
    # Поиск выгодных
    async def searchBenefit(self, excelInfo, dialog):
        autooptGoods = []

        # Получение индикатора загрузки
        progressBar = dialog.ids["content_container"].children[0].children[0] 

        for goods in excelInfo["goods"]:
            Animation(value=(excelInfo["goods"].index(goods) * 100) / len(excelInfo["goods"]), duration=1.).start(progressBar)
            exel_count = goods["amount"]

            autopiterOffers = self.get_info_autopiter
            # Поиск в автоальянсе по артиколу
            autooptOffers = self.get_product_info

            all_search = await asyncio.gather(*[
                autopiterOffers((goods["mainArtikul"])), 
                autooptOffers(goods["mainArtikul"], goods["name"])
            ])

            autopiterOffers = all_search[0]
            autooptOffers = all_search[1]
            print("ALL_SEARCH", all_search)
            best_offer = {
                "Название": goods["name"],
                "Название из экселя": goods["name"],
                "Артикул": goods["mainArtikul"],
                "Артикул из экселя": goods["mainArtikul"],
                "Цена": 0,
                "Количество": 0,
                "Количество из экселя": goods["amount"],
                "Источник": "АвтоАльянс",
                "Бренд": "",
            }
            current_price = 1000000000.00
            sub_current_price = 0

            for offer in autooptOffers:

                if type(offer['Цены товара']) == dict:

                    if type(offer['Количество на складе']) == int:
                        site_count = int(offer['Количество на складе'])
                    else:
                        site_count = 0

                    sub_current_price = offer['Цены товара']['Опт 3']
                    if int(exel_count) <= int(site_count):
                        
                        if float(offer['Цены товара']['Опт 3']) < float(current_price):
                            current_price = float(offer['Цены товара']['Опт 3'])

                            best_offer["Количество"] = site_count
                            best_offer["Количество из экселя"] = goods['amount']
                            best_offer["Цена"] = current_price
                            best_offer["Артикул"] = offer["Артикул товара"]
                            best_offer["Артикул из экселя"] = goods["mainArtikul"]
                            best_offer["Бренд"] = offer["Бренд товара"]
                            best_offer["Название"] = offer["Название"]
                            best_offer["Название из экселя"] = goods["name"]
                else:

                    for analog_offer in offer['Цены товара']:

                        if type(analog_offer['Количество']) == int:
                            site_count = int(analog_offer['Количество'])
                        else:
                            site_count = 0

                        sub_current_price = analog_offer['Цена']
                        if int(exel_count) <= int(site_count):
                            
                            if float(analog_offer['Цена']) < float(current_price):
                                current_price = float(analog_offer['Цена'])

                                best_offer["Количество"] = site_count
                                best_offer["Количество из экселя"] = goods['amount']
                                best_offer["Цена"] = current_price
                                best_offer["Артикул"] = offer["Артикул товара"]
                                best_offer["Артикул из экселя"] = goods["mainArtikul"]
                                best_offer["Бренд"] = offer["Бренд товара"]
                                best_offer["Название"] = offer["Название"]
                                best_offer["Название из экселя"] = goods["name"]

                # print(type(offer['Цены товара']) == dict)
            if len(autooptOffers) != 0 and best_offer["Количество"] == 0:
                best_offer["Цена"] = sub_current_price
                best_offer["Количество"] = site_count
            elif len(autooptOffers) == 0:
                best_offer["Количество"] = "Товар не найден"
                best_offer["Цена"] = "Товар не найден"
            autooptGoods.append(best_offer)
            sleep_time = uniform(5, 7)
            sleep(sleep_time)

        progressBar.value = 0
        return autooptGoods

    

if __name__ == "__main__":
   app = EngineParser()
   app.run()